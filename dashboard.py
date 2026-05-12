from flask import Flask, render_template_string, request, send_from_directory, redirect, send_file
import csv, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

app = Flask(__name__)

HTML = """
<!DOCTYPE html>
<html>
<head>
<title>AI PCB Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>

<style>
body{
margin:0;
font-family:Segoe UI;
background:#020617;
color:white;
}

.container{
max-width:1100px;
margin:auto;
padding:20px;
}

.header{
text-align:center;
margin-bottom:10px;
}

.header h2{
background:linear-gradient(90deg,#38bdf8,#22c55e);
-webkit-background-clip:text;
color:transparent;
font-size:28px;
}

.time{
color:#94a3b8;
font-size:14px;
}

a{
text-decoration:none;
color:inherit;
}

/* 🔥 STATUS GLOW */
.status{
padding:12px;
text-align:center;
border-radius:8px;
font-weight:bold;
margin:20px 0;
animation: glow 1.5s infinite alternate;
}

@keyframes glow{
from{box-shadow:0 0 10px rgba(0,255,0,0.3);}
to{box-shadow:0 0 20px rgba(0,255,0,0.7);}
}

.stats{
display:flex;
justify-content:center;
gap:15px;
margin-bottom:25px;
}

.card{
background:#0f172a;
padding:15px;
border-radius:10px;
width:120px;
text-align:center;
transition:0.3s;
border:2px solid transparent;
cursor:pointer;
}

/* 🔥 ACTIVE CARD */
.card.active{
background:#1e293b;
box-shadow:0 0 15px rgba(56,189,248,0.6);
}

.card.total{border-color:#38bdf8;}
.card.pass{border-color:#22c55e;}
.card.fail{border-color:#ef4444;}
.card.rate{border-color:#06b6d4;}

.card:hover{
transform:translateY(-5px) scale(1.05);
box-shadow:0 0 15px rgba(56,189,248,0.4);
}

.controls{
display:flex;
justify-content:center;
margin-bottom:20px;
gap:10px;
}

.search-box{
display:flex;
gap:8px;
}

.search-box input{
padding:8px 12px;
border-radius:8px;
border:none;
background:#0f172a;
color:white;
width:200px;
outline:none;
}

/* 🔥 SEARCH GLOW */
.search-box input:focus{
box-shadow:0 0 10px #38bdf8;
}

.btn{
padding:8px 12px;
border:none;
border-radius:8px;
cursor:pointer;
transition:0.3s;
}

.btn:hover{
transform:scale(1.05);
box-shadow:0 0 10px rgba(255,255,255,0.2);
}

.search-btn{background:#1e293b;color:white;}
.search-btn:hover{background:#38bdf8;}

.reset-btn{background:#ef4444;color:white;}
.reset-btn:hover{background:#dc2626;}

.export-btn{background:#22c55e;color:black;}
.export-btn:hover{background:#16a34a;}

.grid{
display:grid;
grid-template-columns:2fr 1fr;
gap:20px;
}

.box{
background:#0f172a;
padding:15px;
border-radius:12px;
}

.preview{
width:100%;
border-radius:10px;
cursor:pointer;
}

table{
width:100%;
border-collapse:collapse;
margin-top:20px;
}

th,td{
padding:10px;
border-bottom:1px solid rgba(255,255,255,0.1);
text-align:center;
}

.row{
transition:0.3s;
}

.row:hover{
background:#1e293b;
transform:scale(1.01);
box-shadow:0 0 10px rgba(56,189,248,0.3);
}

.badge{
padding:5px 10px;
border-radius:6px;
font-size:12px;
font-weight:bold;
}

.badge.pass{
background:#22c55e;
color:black;
}

.chips{
display:flex;
flex-wrap:wrap;
gap:5px;
justify-content:center;
}

.chip{
background:#ef4444;
padding:4px 8px;
border-radius:6px;
font-size:11px;
}

.table-img{
width:60px;
cursor:pointer;
transition:0.3s;
}

.table-img:hover{
transform:scale(1.3);
}

#preview-box{
position:fixed;
top:60px;
right:20px;
display:none;
z-index:999;
}

#preview-box img{
width:250px;
border-radius:10px;
box-shadow:0 0 20px black;
}

.modal{
display:none;
position:fixed;
top:0;
left:0;
width:100%;
height:100%;
background:rgba(0,0,0,0.95);
}

.modal-content{
width:80%;
margin:auto;
margin-top:3%;
text-align:center;
}

.modal img{
width:100%;
max-height:80vh;
object-fit:contain;
cursor:grab;
}
</style>
</head>

<body>

<div class="container">

<div class="header">
<h2>AI PCB Inspection System</h2>
<div class="time">{{date}} | {{time}}</div>
</div>

<div class="status" style="background: {{ 'green' if last=='NONE' else 'red' }}">
{{ 'PASS' if last=='NONE' else 'FAIL' }} • 
{{ 0 if last=='NONE' else last.count(",")+1 }} Defects
</div>

<div class="stats">
<a href="/?filter=all"><div class="card total {% if filter=='all' %}active{% endif %}">Total<br>{{total}}</div></a>
<a href="/?filter=pass"><div class="card pass {% if filter=='pass' %}active{% endif %}">Passed<br>{{passed}}</div></a>
<a href="/?filter=fail"><div class="card fail {% if filter=='fail' %}active{% endif %}">Failed<br>{{failed}}</div></a>
<div class="card rate">Rate<br>{{percent}}%</div>
</div>

<div class="controls">
<form method="get" class="search-box">
<input name="search" placeholder="Search by Board ID...">
<button class="btn search-btn">Search</button>
</form>

<form action="/reset" method="post">
<button class="btn reset-btn">Reset</button>
</form>

<a href="/download">
<button class="btn export-btn">Export Excel</button>
</a>
</div>

<div class="grid">

<div class="box">
<h3>Latest Board (ID: {{latest_id}})</h3>
{% if latest %}
<img src="/{{latest}}" class="preview"
onclick="openModal('{{latest}}','{{last}}','{{latest_id}}')">
{% endif %}
</div>

<div class="box">
<h3 style="text-align:center;">Summary</h3>
<div style="display:flex;justify-content:center;">
<canvas id="pie" style="width:240px;height:240px;"></canvas>
</div>
</div>

</div>

<table>
<tr>
<th>ID</th>
<th>Status</th>
<th>Defects</th>
<th>Time</th>
<th>Image</th>
</tr>

{% for r in data %}
<tr class="row">

<td>#{{ "%04d"|format(r[0]|int) }}</td>

<td>
{% if r[1]=='NONE' %}
<span class="badge pass">PASS</span>
{% else %}
<div class="chips">
{% for item in r[1].split(",") %}
<span class="chip">{{item}}</span>
{% endfor %}
</div>
{% endif %}
</td>

<td>
{% if r[1]=='NONE' %}
0
{% else %}
{{ r[1].count(",") + 1 }}
{% endif %}
</td>

<td>{{r[2]}}</td>

<td>
<img src="/{{r[3]}}" class="table-img"
onmouseover="preview(this.src)"
onclick="openModal('{{r[3]}}','{{r[1]}}','{{r[0]}}')">
</td>

</tr>
{% endfor %}
</table>

</div>

<div id="preview-box"></div>

<div id="modal" class="modal" onclick="closeModal()">
<div class="modal-content" onclick="event.stopPropagation()">
<h2 id="title"></h2>
<img id="img">
<p id="text"></p>
</div>
</div>

<script>
setTimeout(()=>location.reload(),5000);

// 🔥 FAIL SOUND
if("{{last}}" !== "NONE"){
new Audio("https://www.soundjay.com/buttons/beep-01a.mp3").play();
}

function preview(src){
let box=document.getElementById("preview-box");
box.innerHTML="<img src='"+src+"'>";
box.style.display="block";
}

document.addEventListener("mouseout",()=>{
document.getElementById("preview-box").style.display="none";
});

let scale=1,x=0,y=0,drag=false,sx,sy;

function openModal(img,status,id){
const image=document.getElementById("img");
document.getElementById("modal").style.display="block";
image.src=img;
document.getElementById("text").innerHTML=status;
document.getElementById("title").innerHTML="Board "+id;
scale=1;x=0;y=0;
image.style.transform="scale(1)";
}

function closeModal(){
document.getElementById("modal").style.display="none";
}

document.getElementById("img").addEventListener("wheel",function(e){
e.preventDefault();
scale+=e.deltaY*-0.001;
scale=Math.min(Math.max(1,scale),5);
this.style.transform=`scale(${scale}) translate(${x}px,${y}px)`;
});

document.getElementById("img").addEventListener("mousedown",function(e){
drag=true;
sx=e.clientX-x;
sy=e.clientY-y;
});

document.addEventListener("mousemove",function(e){
if(!drag)return;
x=e.clientX-sx;
y=e.clientY-sy;
document.getElementById("img").style.transform=`scale(${scale}) translate(${x}px,${y}px)`;
});

document.addEventListener("mouseup",()=>drag=false);

// 🔥 SMOOTH PIE
new Chart(document.getElementById("pie"),{
type:"doughnut",
data:{
labels:["Pass","Fail"],
datasets:[{
data:[{{passed}},{{failed}}],
backgroundColor:["#22c55e","#ef4444"],
borderWidth:2,
hoverOffset:15
}]
},
options:{
animation:{duration:1200},
cutout:"65%",
plugins:{legend:{labels:{color:"white"}}}
}
});
</script>

</body>
</html>
"""

@app.route("/")
def home():
    filter_type=request.args.get("filter","all")
    search=request.args.get("search","")

    data=[]
    if os.path.exists("data/log.csv"):
        with open("data/log.csv") as f:
            reader=csv.reader(f)
            next(reader)
            for r in reader:
                if search and search != r[0]:
                    continue
                if filter_type=="pass" and r[1]!="NONE":
                    continue
                if filter_type=="fail" and r[1]=="NONE":
                    continue
                data.append(r)

    total=len(data)
    passed=sum(1 for r in data if r[1]=="NONE")
    failed=total-passed
    percent=round((passed/total)*100,2) if total else 0

    latest=data[-1][3] if data else ""
    latest_id=data[-1][0] if data else "-"
    last=data[-1][1] if data else "NONE"

    now=datetime.now()

    return render_template_string(HTML,
        data=data[::-1],
        total=total,
        passed=passed,
        failed=failed,
        percent=percent,
        latest=latest,
        latest_id=latest_id,
        last=last,
        filter=filter_type,
        time=now.strftime("%H:%M:%S"),
        date=now.strftime("%Y-%m-%d")
    )

@app.route("/download")
def download():
    wb = Workbook()
    ws = wb.active
    ws.title = "PCB Report"

    headers = ["ID", "Status", "Defect Count", "Time", "Image"]
    ws.append(headers)

    row_index = 2

    if os.path.exists("data/log.csv"):
        with open("data/log.csv") as f:
            reader = csv.reader(f)
            next(reader)
            for r in reader:

                status = r[1]
                defect_count = 0 if status == "NONE" else status.count(",") + 1

                ws.cell(row=row_index, column=1, value=f"#{int(r[0]):04d}")
                ws.cell(row=row_index, column=2, value=status)
                ws.cell(row=row_index, column=3, value=defect_count)
                ws.cell(row=row_index, column=4, value=r[2])

                if os.path.exists(r[3]):
                    try:
                        img = XLImage(r[3])
                        img.width = 80
                        img.height = 60
                        ws.add_image(img, f"E{row_index}")
                        ws.row_dimensions[row_index].height = 50
                    except:
                        pass

                row_index += 1

    filename = "PCB_Report.xlsx"
    wb.save(filename)

    return send_file(filename, as_attachment=True)

@app.route("/reset", methods=["POST"])
def reset():
    if os.path.exists("data/log.csv"):
        with open("data/log.csv","w",newline="") as f:
            writer=csv.writer(f)
            writer.writerow(["ID","Status","Time","Image"])
    return redirect("/")

@app.route('/data/images/<path:filename>')
def images(filename):
    return send_from_directory('data/images', filename)

app.run(debug=True)